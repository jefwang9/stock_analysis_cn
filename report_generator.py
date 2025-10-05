"""
报表生成模块
生成每日预测报告和准确率分析报告
"""
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from typing import List, Dict, Any, Optional
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import logging
from config import settings

logger = logging.getLogger(__name__)

# 设置中文字体
plt.rcParams['font.sans-serif'] = ['SimHei', 'Arial Unicode MS']
plt.rcParams['axes.unicode_minus'] = False

class ReportGenerator:
    """报表生成器"""
    
    def __init__(self):
        self.reports_dir = f"{settings.data_dir}/reports"
        self._ensure_directories()
    
    def _ensure_directories(self):
        """确保报表目录存在"""
        import os
        os.makedirs(self.reports_dir, exist_ok=True)
    
    def generate_daily_prediction_report(self, predictions_df: pd.DataFrame, 
                                       accuracy_stats: Dict[str, Any]) -> str:
        """
        生成每日预测报告
        
        Args:
            predictions_df: 预测结果数据框
            accuracy_stats: 准确率统计
            
        Returns:
            str: 报表文件路径
        """
        # 创建Excel工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "每日预测报告"
        
        # 设置表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 主标题
        ws['A1'] = f"A股板块预测报告 - {datetime.now().strftime('%Y年%m月%d日')}"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:F1')
        
        # 准确率总结
        row = 3
        ws[f'A{row}'] = "预测准确率统计"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws[f'A{row}'].alignment = header_alignment
        
        accuracy_summary = [
            ["指标", "数值"],
            ["总体准确率", f"{accuracy_stats.get('accuracy_rate', 0):.2%}"],
            ["平均置信度", f"{accuracy_stats.get('avg_confidence', 0):.2%}"],
            ["上涨板块准确率", f"{accuracy_stats.get('top_gainer_accuracy', 0):.2%}"],
            ["下跌板块准确率", f"{accuracy_stats.get('top_loser_accuracy', 0):.2%}"],
            ["预测总数", accuracy_stats.get('total_predictions', 0)]
        ]
        
        for i, row_data in enumerate(accuracy_summary):
            ws[f'A{row + 2 + i}'] = row_data[0]
            ws[f'B{row + 2 + i}'] = row_data[1]
        
        # 预测结果详情
        detail_start_row = row + len(accuracy_summary) + 3
        ws[f'A{detail_start_row}'] = "预测结果详情"
        ws[f'A{detail_start_row}'].font = header_font
        ws[f'A{detail_start_row}'].fill = header_fill
        ws[f'A{detail_start_row}'].alignment = header_alignment
        
        # 设置详情表头
        headers = ["板块名称", "预测涨跌幅(%)", "置信度", "排名", "分类"]
        
        detail_header_row = detail_start_row + 1
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=detail_header_row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 填充预测数据
        detail_data_row = detail_header_row + 1
        
        # 排序预测结果
        top_gainers = predictions_df.nlargest(settings.trading.predict_top_n, 'predicted_change')
        top_losers = predictions_df.nsmallest(settings.trading.predict_bottom_n, 'predicted_change')
        
        # 添加top gainers
        for i, (_, row_data) in enumerate(top_gainers.iterrows()):
            row_num = detail_data_row + i
            ws[f'A{row_num}'] = row_data['sector']
            ws[f'B{row_num}'] = f"{row_data['predicted_change']:.2f}"
            ws[f'C{row_num}'] = f"{row_data.get('confidence', 0):.2%}"
            ws[f'D{row_num}'] = i + 1
            ws[f'E{row_num}'] = "预测上涨"
            
            # 设置红色背景
            ws[f'E{row_num}'].fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        
        # 添加top losers
        loser_start_row = detail_data_row + len(top_gainers)
        for i, (_, row_data) in enumerate(top_losers.iterrows()):
            row_num = loser_start_row + i
            ws[f'A{row_num}'] = row_data['sector']
            ws[f'B{row_num}'] = f"{row_data['predicted_change']:.2f}"
            ws[f'C{row_num}'] = f"{row_data.get('confidence', 0):.2%}"
            ws[f'D{row_num}'] = -(i + 1)
            ws[f'E{row_num}'] = "预测下跌"
            
            # 设置绿色背景
            ws[f'E{row_num}'].fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        
        # 调整列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 12
        
        # 保存文件
        filename = f"daily_prediction_report_{datetime.now().strftime('%Y%m%d')}.xlsx"
        filepath = f"{self.reports_dir}/{filename}"
        wb.save(filepath)
        
        logger.info(f"每日预测报告已生成: {filepath}")
        
        return filepath
    
    def generate_accuracy_report(self, period_days: int = 30) -> str:
        """
        生成准确率分析报告
        
        Args:
            period_days: 统计周期天数
            
        Returns:
            str: 报表文件路径
        """
        # 导入backtesting模块
        from backtesting import Backtester
        
        backtester = Backtester()
        performance_data = backtester.get_performance_report(period_days)
        
        if performance_data.empty:
            logger.warning("没有足够的数据生成准确率报告")
            return ""
        
        # 创建Excel工作簿
        wb = Workbook()
        
        # 1. 总体统计表页
        ws_summary = wb.active
        ws_summary.title = "总体统计"
        
        # 主标题
        ws_summary['A1'] = f"A股预测准确率分析报告 - 最近{period_days}天"
        ws_summary['A1'].font = Font(bold=True, size=16)
        ws_summary.merge_cells('A1:F1')
        
        # 统计摘要
        row = 3
        summary_stats = [
            ["统计指标", "数值"],
            ["统计天数", f"{len(performance_data)} 天"],
            ["平均准确率", f"{performance_data['accuracy_rate'].mean():.2%}"],
            ["最高准确率", f"{performance_data['accuracy_rate'].max():.2%}"],
            ["最低准确率", f"{performance_data['accuracy_rate'].min():.2%}"],
            ["标准差", f"{performance_data['accuracy_rate'].std():.2%}"],
            ["平均置信度", f"{performance_data['avg_confidence'].mean():.2%}"],
            ["上涨板块平均准确率", f"{performance_data['top_gainer_accuracy'].mean():.2%}"],
            ["下跌板块平均准确率", f"{performance_data['top_loser_accuracy'].mean():.2%}"]
        ]
        
        for i, row_data in enumerate(summary_stats):
            ws_summary[f'A{row + i}'] = row_data[0]
            ws_summary[f'B{row + i}'] = row_data[1]
        
        # 2. 详细数据表页
        ws_detail = wb.create_sheet("详细数据")
        
        # 设置表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 表头
        headers = ["日期", "预测总数", "正确预测数", "准确率", "平均置信度", 
                 "上涨板块准确率", "下跌板块准确率"]
        
        for col, header in enumerate(headers, 1):
            cell = ws_detail.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 填充数据
        for i, row_data in enumerate(performance_data.itertuples(), 2):
            ws_detail[f'A{i}'] = row_data.date
            ws_detail[f'B{i}'] = row_data.total_predictions
            ws_detail[f'C{i}'] = row_data.correct_predictions
            ws_detail[f'D{i}'] = f"{row_data.accuracy_rate:.2%}"
            ws_detail[f'E{i}'] = f"{row_data.avg_confidence:.2%}"
            ws_detail[f'F{i}'] = f"{row_data.top_gainer_accuracy:.2%}"
            ws_detail[f'G{i}'] = f"{row_data.top_loser_accuracy:.2%}"
        
        # 3. 趋势分析图表
        self._create_trend_charts(performance_data)
        
        # 保存文件
        filename = f"accuracy_report_{period_days}days_{datetime.now().strftime('%Y%m%d')}.xlsx"
        filepath = f"{self.reports_dir}/{filename}"
        wb.save(filepath)
        
        logger.info(f"准确率分析报告已生成: {filepath}")
        
        return filepath
    
    def _create_trend_charts(self, performance_data: pd.DataFrame):
        """创建趋势分析图表"""
        if performance_data.empty:
            return
        
        # 创建图表
        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2, figsize=(15, 10))
        fig.suptitle(f'A股预测模型性能分析 - 最近{len(performance_data)}天', fontsize=16)
        
        # 1. 准确率趋势
        ax1.plot(performance_data['date'], performance_data['accuracy_rate'], 
                marker='o', linewidth=2, color='blue')
        ax1.set_title('预测准确率趋势')
        ax1.set_ylabel('准确率')
        ax1.grid(True, alpha=0.3)
        ax1.set_ylim(0, 1)
        
        # 2. 置信度趋势
        ax2.plot(performance_data['date'], performance_data['avg_confidence'], 
                marker='s', linewidth=2, color='green')
        ax2.set_title('平均置信度趋势')
        ax2.set_ylabel('置信度')
        ax2.grid(True, alpha=0.3)
        ax2.set_ylim(0, 1)
        
        # 3. 上涨vs下跌板块准确率对比
        ax3.plot(performance_data['date'], performance_data['top_gainer_accuracy'], 
                marker='o', linewidth=2, color='red', label='上涨板块')
        ax3.plot(performance_data['date'], performance_data['top_loser_accuracy'], 
                marker='s', linewidth=2, color='green', label='下跌板块')
        ax3.set_title('上涨vs下跌板块预测准确率')
        ax3.set_ylabel('准确率')
        ax3.legend()
        ax3.grid(True, alpha=0.3)
        ax3.set_ylim(0, 1)
        
        # 4. 累计准确率
        cumulative_acc = performance_data['accuracy_rate'].expanding().mean()
        ax4.plot(performance_data['date'], cumulative_acc, 
                marker='o', linewidth=2, color='purple')
        ax4.set_title('累计平均准确率')
        ax4.set_ylabel('累计准确率')
        ax4.grid(True, alpha=0.3)
        ax4.set_ylim(0, 1)
        
        # 调整布局
        plt.tight_layout()
        plt.xticks(rotation=45)
        
        # 保存图表
        chart_path = f"{self.reports_dir}/performance_trend_{datetime.now().strftime('%Y%m%d')}.png"
        plt.savefig(chart_path, dpi=300, bbox_inches='tight')
        plt.close()
        
        logger.info(f"趋势分析图表已保存: {chart_path}")
    
    def generate_sector_analysis_report(self, sector_performance: Dict[str, Any]) -> str:
        """
        生成板块分析报告
        
        Args:
            sector_performance: 板块性能数据
            
        Returns:
            str: 报表文件路径
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "板块分析报告"
        
        # 主标题
        ws['A1'] = f"板块预测分析报告 - {sector_performance.get('sector', '未知板块')}"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:F1')
        
        # 统计周期
        row = 3
        ws[f'A{row}'] = "分析周期"
        ws[f'B{row}'] = sector_performance.get('period', '')
        
        # 板块表现统计
        row += 2
        stats_items = [
            ["总预测次数", sector_performance.get('total_predictions', 0)],
            ["准确率", f"{sector_performance.get('accuracy_rate', 0):.2%}"],
            ["上涨预测准确率", f"{sector_performance.get('positive_accuracy', 0):.2%}"],
            ["下跌预测准确率", f"{sector_performance.get('negative_accuracy', 0):.2%}"],
            ["平均绝对误差", f"{sector_performance.get('mae', 0):.2%}"],
            ["均方根误差", f"{sector_performance.get('rmse', 0):.2%}"],
            ["平均预测涨跌幅", f"{sector_performance.get('avg_predicted_change', 0):.2%}"],
            ["实际平均涨跌幅", f"{sector_performance.get('actual_change', 0):.2%}"],
            ["平均置信度", f"{sector_performance.get('confidence', 0):.2%}"]
        ]
        
        for stat_item in stats_items:
            ws[f'A{row}'] = stat_item[0]
            ws[f'B{row}'] = stat_item[1]
            row += 1
        
        # 调整列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        
        # 保存文件
        sector_name = sector_performance.get('sector', 'unknown').replace('/', '_')
        filename = f"Sector_Analysis_Report_{sector_name}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        filepath = f"{self.reports_dir}/{filename}"
        wb.save(filepath)
        
        logger.info(f"板块分析报告已生成: {filepath}")
        
        return filepath
    
    def generate_comprehensive_report(self, predictions_df: pd.DataFrame, 
                                    accuracy_stats: Dict[str, Any], 
                                    period_days: int = 30) -> Dict[str, str]:
        """
        生成综合报告包
        
        Args:
            predictions_df: 预测结果
            accuracy_stats: 准确率统计
            period_days: 统计周期
            
        Returns:
            Dict: 所有生成报告的文件路径
        """
        reports = {}
        
        # 1. 每日预测报告
        reports['daily_prediction'] = self.generate_daily_prediction_report(
            predictions_df, accuracy_stats)
        
        # 2. 准确率分析报告
        reports['accuracy_analysis'] = self.generate_accuracy_report(period_days)
        
        # 3. 获取所有板块的详细分析
        from backtesting import Backtester
        
        backtester = Backtester()
        
        # 为每个主要板块生成分析报告
        sector_reports = {}
        for sector in settings.trading.sectors[:10]:  # 限制前10个板块
            sector_performance = backtester.get_sector_performance_report(sector, period_days)
            if 'error' not in sector_performance:
                sector_file = self.generate_sector_analysis_report(sector_performance)
                sector_reports[sector] = sector_file
        
        reports['sector_analysis'] = sector_reports
        
        logger.info(f"综合报告包生成完成，包含 {len(reports)} 类报告")
        
        return reports

# 便捷函数
def generate_today_report(predictions_df: pd.DataFrame, 
                        accuracy_stats: Dict[str, Any]) -> str:
    """生成今日报告"""
    generator = ReportGenerator()
    return generator.generate_daily_prediction_report(predictions_df, accuracy_stats)

def generate_weekly_report(period_days: int = 7) -> str:
    """生成周报"""
    generator = ReportGenerator()
    return generator.generate_accuracy_report(period_days)

def generate_monthly_report(period_days: int = 30) -> str:
    """生成月报"""
    generator = ReportGenerator()
    return generator.generate_accuracy_report(period_days)
